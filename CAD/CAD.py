from tkinter import *
from datetime import date
from datetime import datetime
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os 
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from tkinter import Tk, Label, PhotoImage 
from tkcalendar import Calendar, DateEntry 
from tkinter import font



filename = None 
background="#F3F7EC"
framebg="#fdffc1"
framefg="#06283D"

root=Tk()
root.title("Auto-CAD Registration")
root.geometry("1200x600")
root.config(bg=background)

# Set window size (resolution: 1200x600)
window_width = 1200
window_height = 600
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate center position
position_top = int(screen_height / 2 - window_height / 2)
position_left = int(screen_width / 2 - window_width / 2)

# Set the window geometry with center position
root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

file=pathlib.Path('CAD.xlsx')
if file.exists() :
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No. "
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Cast"
    sheet['H1']="Contact"
    sheet['I1']="Total Fee"
    sheet['J1']="Paid Fee"
    sheet['K1']="Remaining Fee"

    file.save('CAD.xlsx')


    
#Exit window
def Exit():
     root.destroy()

def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file", 
                                          filetypes=(("JPG File", "*.jpg"), ("PNG File", "*.png"), ("All files", "*.*")))
    
    # Check if a file was selected
    if filename:
        try:
            img = Image.open(filename)
            resized_image = img.resize((190, 190))  # Resize the image to fit the label
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load image: {e}")
    else:
        messagebox.showwarning("No File Selected", "Please select a valid image file.")


#create automatic registration no. 
def registration_no():
    file=openpyxl.load_workbook('CAD.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        registration.set(max_row_value+1)
    except:
        registration.set("1")    

#Contact input limit
def validate_contact_input(value_if_allowed):
    if value_if_allowed.isdigit() and len(value_if_allowed) <= 10:
        return True
    elif value_if_allowed == "":  # Allow clearing the input
        return True
    else:
        return False

vcmd = root.register(validate_contact_input)

#Dob date validation
def validate_dob_format(dob):
    try:
        dob_date = datetime.strptime(dob, "%d/%m/%Y")
        today = datetime.today()
        age = (today - dob_date).days // 365
        if 10 <= age <= 80:
            return True
        else:
            messagebox.showerror("Invalid DOB", "Age must be between 10 and 80 years.")
            return False
    except ValueError:
        messagebox.showerror("Invalid DOB", "Please enter DOB in the format dd/mm/yyyy.")
        return False

    

#Remaning fee generator
def calculate_remaining_fee(event):
    try:
        paid_fee = Paid_fee.get()  
        total_fee = Total_fee.get()
        remaining_fee = total_fee - paid_fee  
        Remaining_fee.set(remaining_fee) 
    except ValueError:
        messagebox.showerror("Invalid input", "Please enter a valid paid fee value.")


#Function to validate the name input (only alphabets and spaces allowed)
def validate_name_input(value_if_allowed):
    if value_if_allowed.replace(" ", "").isalpha() or value_if_allowed == "":  
        return True
    else:
        return False
vcmd_name = root.register(validate_name_input)  # Register the validation command for Name


#Function to validate the DOB input (only numbers, spaces, and hyphens allowed)
def validate_dob_input(value_if_allowed):
    if value_if_allowed.replace(" ", "").replace("-" , "").replace("/", "").isdigit() or value_if_allowed == "":  # Allow numbers, spaces, and hyphens
        return True
    else:
        return False
vcmd_dob = root.register(validate_dob_input) # Register the validation command for DOB

#clear
def Clear():
    global img
    Name.set('')  
    DOB.set('')  
    Cast.set("Select Cast") 
    Contact.set('')  
    Paid_fee.set('')
    Total_fee.set('2500') 
    Remaining_fee.set('')
    Qualification.set("Select Class")
    registration_no() 
    saveButton.config(state='normal') 

    name_entry.config(bg="white")
    
    img1 = PhotoImage(file='Images/profile.png')
    lbl.config(image=img1)
    lbl.image = img1

    img = "" 

#SAVE 
def Save():

    
    R1 = registration.get()
    N1 = Name.get()
    C1 = Qualification.get()
    selection()
    try:
        G1 = gender 
    except:
        messagebox.showerror("Error", "Please select a gender!")
        return

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Cast.get()
    S1 = Contact.get()
    Paidfee = Paid_fee.get()
    T1 = Total_fee.get()
    r1 = Remaining_fee.get()

    if not validate_dob_format(DOB.get()):
        return  # Stop further execution if DOB is invalid
    
     
    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "Select Cast" or S1 == "" or Paidfee == "" or T1 == "" or r1 == "":
        messagebox.showerror("Error", "Some required fields are missing!")
        return


    if not filename:  
        messagebox.showwarning("No Image", "Please upload an image before saving.")
        return

    try:

        file = openpyxl.load_workbook('CAD.xlsx')
        sheet = file.active

        next_row = sheet.max_row + 1
        
        sheet.append([R1, N1, C1, G1, D2, D1, Re1, S1, T1, Paidfee, r1])

        file.save('CAD.xlsx')

        try:
            img.save("Student Images/" + str(R1) + ".jpg")
        except:
            messagebox.showwarning("Image Not Saved", "Profile image could not be saved.")

        messagebox.showinfo("Success", "Data entered successfully!")
        Clear() 
        registration_no() 

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving the data: {e}")

        messagebox.showinfo("Success", "Data entered successfully!")
        Clear() 
        registration_no()

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving the data: {e}")

        
#Search
def search():
    text=Search.get()
    Clear()
    saveButton.config(state='disable')
    file=openpyxl.load_workbook("CAD.xlsx")
    sheet=file.active

    for row in sheet.rows:
         if row[0].value==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number!!")

    #reg_no_position showing like A1,A2,A3......An
    #but reg_number just shoing number afer A2 like 2,3.......,n

    x1=sheet.cell(row=int(reg_number),column=1).value   
    x2=sheet.cell(row=int(reg_number),column=2).value  
    x3=sheet.cell(row=int(reg_number),column=3).value  
    x4=sheet.cell(row=int(reg_number),column=4).value  
    x5=sheet.cell(row=int(reg_number),column=5).value  
    x6=sheet.cell(row=int(reg_number),column=6).value  
    x7=sheet.cell(row=int(reg_number),column=7).value  
    x8=sheet.cell(row=int(reg_number),column=8).value  
    x9=sheet.cell(row=int(reg_number),column=9).value  
    x10=sheet.cell(row=int(reg_number),column=10).value  
    x11=sheet.cell(row=int(reg_number),column=11).value  

    registration.set(x1)
    Name.set(x2) 
    Qualification.set(x3)

    if x4=='Female':
        R2.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Cast.set(x7)
    Contact.set(x8)
    Total_fee.set(x9)
    Paid_fee.set(x10)
    Remaining_fee.set(x11)

    img=(Image.open("student images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#update  
def update():
     R1=registration.get()
     N1=Name.get()
     C1=Qualification.get()
     selection()
     G1=gender

     D2=DOB.get()
     D1=Date.get()
     Re1=Cast.get()
     S1=Contact.get()
     Paidfee=Paid_fee.get()
     T1=Total_fee.get()
     r1=Remaining_fee.get()

     file=openpyxl.load_workbook("CAD.xlsx")
     sheet=file.active

     for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            print(reg_number)

     sheet.cell(column=1,row=int(reg_number),value=R1)
     sheet.cell(column=2,row=int(reg_number),value=N1)
     sheet.cell(column=3,row=int(reg_number),value=C1)
     sheet.cell(column=4,row=int(reg_number),value=G1)
     sheet.cell(column=5,row=int(reg_number),value=D2)
     sheet.cell(column=6,row=int(reg_number),value=D1)
     sheet.cell(column=7,row=int(reg_number),value=Re1)
     sheet.cell(column=8,row=int(reg_number),value=S1)
     sheet.cell(column=9,row=int(reg_number),value=T1)
     sheet.cell(column=10,row=int(reg_number),value=Paidfee)
     sheet.cell(column=11,row=int(reg_number),value=r1)
        
     file.save(r'CAD.xlsx')

     try:
        img.save("student images/"+str(R1)+".jpg")
     except:
        pass

     messagebox.showinfo("Update","Updated Sucessfully!!")

     Clear()


#gender 
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"


#Background image
frame=Frame(root,bg="red")
frame.pack(fill=Y)


backgroundimage=PhotoImage(file="Images/BG.png")
Label(frame,image=backgroundimage).pack()

#top frames
Label(root,text="New Student Registration for Auto-CAD", font='Palatino 20',anchor='center',fg="#000",bg="#f9fcae").place(x=376,y=35)


#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=14,bd=2,font="arial 14",justify="center").place(x=720,y=128)
srch=Button(root,text="Search",compound=LEFT ,width=10,bg='#ffc31b',font="arial 11 bold",command=search)
srch.place(x=610,y=125)

registration=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=registration,width=12,bd=2,font="arial 14")
reg_entry.place(x=190,y=130)

registration_no() 

today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=12,bd=2,font="arial 14")
date_entry.place(x=418,y=130)

Date.set(d1)

#student details
obj=LabelFrame(root,text="Student's Details",font="arial 12 bold" ,bd=2,width=900,bg=framebg,fg=framefg,height=180,relief=GROOVE)
obj.place(x=30,y=190)

Label(root,text="Full Name: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=40,y=225)
Label(root,text="Date of Birth: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=40,y=275)
Label(root, text="Format: dd/mm/yyyy", font="Bookman 8 italic", bg=framebg, fg=framefg).place(x=265, y=300)
Label(root,text="Gender: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=40,y=325)

Label(root,text="Qualification: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=550,y=225)
Label(root,text="Cast: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=550,y=275)
Label(root,text="Contact: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=550,y=325)

Name = StringVar()
global name_entry
name_entry = Entry(obj, textvariable=Name, width=20, font="Roboto 13", validate="key", validatecommand=(vcmd_name, "%P"))
name_entry.place(x=150, y=17)

DOB=StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font="Roboto 13", validate="key", validatecommand=(vcmd_dob, "%P"))
dob_entry.insert(0, "dd/mm/yyyy")  # Placeholder text
dob_entry.bind("<FocusIn>", lambda event: dob_entry.delete(0, "end"))  # Clear placeholder on focus
dob_entry.place(x=150, y=66)


radio=IntVar()
R1= Radiobutton(obj,text="Male",variable=radio,font="Roboto 12", value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=115)

R2= Radiobutton(obj,text="Female",variable=radio,font="Roboto 12", value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=210,y=115)

Cast=Combobox(obj,values=["OPEN","SC","OBC","ST","NT"],font="Roboto 13",width=17,state="r")
Cast.place(x=650,y=65)
Cast.set("Select Cast")

Contact=IntVar()
contact_entry = Entry(obj, textvariable=Contact, width=20, font="arial 12", 
                      validate="key", validatecommand=(vcmd, "%P"))
contact_entry.place(x=650, y=115)

Qualification=Combobox(obj,values=["1-5","6-10","11-12","Diploma","Degree","Vocational","Other"],font="Roboto 12",width=17,state="r")
Qualification.place(x=650,y=15)
Qualification.set("Select Class")



#Fee details
obj2=LabelFrame(root,text="Fee Details",font="arial 12 bold",bd=2,width=900,bg=framebg,fg=framefg,height=180,relief=GROOVE)
obj2.place(x=30,y=400)

Label(obj2,text="Total fee: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=280,y=30)

Total_fee=IntVar()
Total_fee.set(2500)  # Set default value to 2500

TF_entry=Entry(obj2,textvariable=Total_fee,width=20,font="Roboto 13")
TF_entry.place(x=380,y=30)

Label(obj2,text="Paid fee: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=20,y=100)
Label(obj2,text="Remaining: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=500,y=100)

Paid_fee=IntVar()
paid_entry=Entry(obj2,textvariable=Paid_fee,width=20,font="Roboto 13")
paid_entry.place(x=120,y=100)
paid_entry.bind("<Return>", calculate_remaining_fee)


Remaining_fee=IntVar()
rf_entry=Entry(obj2,textvariable=Remaining_fee,width=20,font="Roboto 13")
rf_entry.place(x=620,y=100)


#image
f=Frame(root,bd=3,bg="black",width=199,height=195,relief=GROOVE)
f.place(x=970,y=150)

img=PhotoImage(file="Images/profile.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button
Button(root,text="Upload",width=16,height=1,font="arial 12 bold",bg="#FF8A8A",command=showimage).place(x=989,y=375)
saveButton=Button(root,text="Save",width=16,height=1,font='arial 13 bold',bg="#c1ff72",command=Save)
saveButton.place(x=989, y=415)
Button(root,text="Reset",width=16,height=1,font="arial 13 bold",bg="#96e1f8",command=Clear).place(x=989,y=455)
Button(root,text="Exit",width=16,height=1,font="arial 13 bold",bg="#ff5c5c",command=Exit).place(x=989,y=535)
Button(root,text="Update",width=16,height=1,font="arial 12 bold",bg="#eba1ef",command=update).place(x=989,y=495)


root.mainloop()
