from tkinter import *
from datetime import date
from datetime import datetime
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os 
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from tkinter import Tk, Label, PhotoImage 
from tkcalendar import Calendar, DateEntry 
from tkinter import font



filename = None 
background="#F3F7EC"
framebg="#fdffc1"
framefg="#06283D"

root=Tk()
root.title("Enquiry Form")
root.geometry("1200x600+75+35")
root.config(bg=background)

# Set window size (resolution: 1200x600)
window_width = 1200
window_height = 600
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate center position
position_top = int(screen_height / 2 - window_height / 2)
position_left = int(screen_width / 2 - window_width / 2)

# Set the window geometry with center position
root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

file=pathlib.Path('Enquiry.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Enquiry No. "
    sheet['B1']="Date of Enquiry"
    sheet['C1']="Name"
    sheet['D1']="Class"
    sheet['E1']="Contact"
    sheet['F1']="Course"
    sheet['G1']="Discription"

    file.save('Enquiry.xlsx')


    
#Exit window
def Exit():
     root.destroy()

#create automatic Enquiry no. 
def Enquiry_no():
    file=openpyxl.load_workbook('Enquiry.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Enquiry.set(max_row_value+1)
    except:
        Enquiry.set("1")    

#Contact input limit
def validate_contact_input(value_if_allowed):
    if value_if_allowed.isdigit() and len(value_if_allowed) <= 10:
        return True
    elif value_if_allowed == "":  # Allow clearing the input
        return True
    else:
        return False

vcmd = root.register(validate_contact_input)

    
#Function to validate the name input (only alphabets and spaces allowed)
def validate_name_input(value_if_allowed):
    if value_if_allowed.replace(" ", "").isalpha() or value_if_allowed == "":  
        return True
    else:
        return False
vcmd_name = root.register(validate_name_input)  # Register the validation command for Name


def Clear():
    Enquiry.set("")  
    Name.set("")  
    Qualification.set("Select Class")  
    Contact.set("")  
    Course.set("Select Course")  
    Enq_entry.delete("1.0", "end")  # Clears Text widget only

    saveButton.config(state='normal')  
    name_entry.config(bg="white")  

    Enquiry_no()  # Regenerate Enquiry No.


    

#SAVE 
def Save():
    R1 = Enquiry.get()
    D1 = Date.get()
    N1 = Name.get()
    C1 = Qualification.get()
    S1 = Contact.get()
    Re1 = Course.get()
    D2 = Enq_entry.get("1.0", "end-1c")  # Get text from Text widget

    if N1 == "" or C1 == "" or D2.strip() == "" or Re1 == "" or S1 == "":
        messagebox.showerror("Error", "Some required fields are missing!")
        return

    try:
        file = openpyxl.load_workbook('Enquiry.xlsx')
        sheet = file.active
        
        # Ensure Enquiry Number is unique
        for row in sheet.iter_rows():
            if row[0].value == R1:
                messagebox.showerror("Error", "Duplicate Enquiry Number!")
                return

        sheet.append([R1, D1, N1, C1, S1, Re1, D2])
        file.save('Enquiry.xlsx')
        file.close()  # Ensure file is saved properly

        messagebox.showinfo("Success", "Data entered successfully!")
        Clear()
        Enquiry_no()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving: {e}")

def search():
    text = Search.get().strip()

    if not text.isdigit():
        messagebox.showerror("Error", "Please enter a valid Enquiry number!")
        return

    Clear()
    saveButton.config(state='disable')

    try:
        file = openpyxl.load_workbook("Enquiry.xlsx")
        sheet = file.active

        for row in sheet.iter_rows():
            if str(row[0].value) == text:  # Match as string
                Enquiry.set(row[0].value)
                Date.set(row[1].value)
                Name.set(row[2].value)
                Qualification.set(row[3].value)
                Contact.set(row[4].value)
                Course.set(row[5].value)

                Enq_entry.delete("1.0", "end")  # Clear previous text
                Enq_entry.insert("1.0", row[6].value)  # Insert new value

                file.close()  # Ensure file is closed after searching
                return

        messagebox.showerror("Error", "Enquiry number not found!")
        file.close()

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")




def update():
    R1 = Enquiry.get()
    D1 = Date.get()
    N1 = Name.get()
    C1 = Qualification.get()
    S1 = Contact.get()
    Re1 = Course.get()
    D2 = Enq_entry.get("1.0", "end-1c")  # Get text from Text widget

    try:
        file = openpyxl.load_workbook("Enquiry.xlsx")
        sheet = file.active
        
        for row in sheet.iter_rows():
            if row[0].value == R1:
                row_num = row[0].row
                sheet.cell(row=row_num, column=1, value=R1)
                sheet.cell(row=row_num, column=2, value=D1)
                sheet.cell(row=row_num, column=3, value=N1)
                sheet.cell(row=row_num, column=4, value=C1)
                sheet.cell(row=row_num, column=5, value=S1)
                sheet.cell(row=row_num, column=6, value=Re1)
                sheet.cell(row=row_num, column=7, value=D2)
                file.save("Enquiry.xlsx")
                messagebox.showinfo("Update", "Updated Successfully!")
                break
        else:
            messagebox.showerror("Error", "Enquiry number not found.")
        Clear()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating: {e}")




#Background image
frame=Frame(root,bg="red")
frame.pack(fill=Y)


backgroundimage=PhotoImage(file="Images/BG.png")
Label(frame,image=backgroundimage).pack()

#top frames
Label(root,text="Admission Inquiry Form", font='Palatino 23',anchor='center',fg="#000",bg="#f9fcae").place(x=460,y=32)


#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=14,bd=2,font="arial 14",justify="center").place(x=720,y=128)
srch=Button(root,text="Search",compound=LEFT ,width=10,bg='#FFD90F',font="arial 12 bold",command=search)
srch.place(x=600,y=125)

Enquiry=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=Enquiry,width=12,bd=2,font="arial 14")
reg_entry.place(x=190,y=130)

Enquiry_no() 

today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=12,bd=2,font="arial 14")
date_entry.place(x=418,y=130)

Date.set(d1)

#student details
obj=LabelFrame(root,text="Student Details",font="arial 12 bold" ,bd=2,width=900,bg=framebg,fg=framefg,height=180,relief=GROOVE)
obj.place(x=30,y=190)

Label(root,text="Full Name: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=40,y=235)
Label(root,text="Qualification: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=550,y=235)
Label(root,text="Course: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=550,y=295)
Label(root,text="Contact: ",font="Bookman 13 bold",bg=framebg,fg=framefg).place(x=40,y=295)

obj2=LabelFrame(root,text="Details of Enquiry",font="arial 12 bold" ,bd=2,width=900,bg=framebg,fg=framefg,height=180,relief=GROOVE)
obj2.place(x=30,y=400)

Enq_entry = Text(root, height=6, width=60, font="arial 14")
Enq_entry.place(x=200,y=420)


Name = StringVar()
global name_entry
name_entry = Entry(obj, textvariable=Name, width=20, font="Roboto 13", validate="key", validatecommand=(vcmd_name, "%P"))
name_entry.place(x=150, y=25)

Course=Combobox(obj,values=["MSCIT","Tally","DTP","Auto-CAD","C","C++","Other"],font="Roboto 13",width=17,state="r")
Course.place(x=650,y=80)


Contact=IntVar()
contact_entry = Entry(obj, textvariable=Contact, width=20,font="arial 12", validate="key", validatecommand=(vcmd, "%P"))
contact_entry.place(x=150, y=85)

Qualification=Combobox(obj,values=["1-5","6-10","11-12","Diploma","Degree","Vocational","Other"],font="Roboto 12",width=17,state="r")
Qualification.place(x=650,y=25)
Qualification.set("Select Class")



#button
saveButton=Button(root,text="Save",width=12,height=3,font='arial 14 bold',bg="#FF9020",command=Save)
saveButton.place(x=989, y=150)
Button(root,text="Reset",width=12,height=3,font="arial 14 bold",bg="#FF9832",command=Clear).place(x=989,y=265)
Button(root,text="Update",width=12,height=3,font="arial 14 bold",bg="#FFB00F",command=update).place(x=989,y=385)
Button(root,text="Exit",width=12,height=3,font="arial 14 bold",bg="#FFD90F",command=Exit).place(x=989,y=500)


root.mainloop()
